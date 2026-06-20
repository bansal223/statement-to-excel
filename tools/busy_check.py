"""Validate a generated statement .xlsx against BUSY import conventions.

Encodes what we learned from a known-good BUSY import file (hdfc_…virk.xlsx):
  - the Date column is plain TEXT in DD/MM/YY(YY) form, not an Excel date cell
  - amount columns hold real numbers, not text
It can't prove BUSY will accept the file (only BUSY can), but it catches the
format regressions that broke the import before.

Usage: python tools/busy_check.py <generated.xlsx> [--date-col Date]
"""
from __future__ import annotations

import re
import sys
from datetime import datetime

from openpyxl import load_workbook

_DATE_TEXT = re.compile(r"^\d{1,2}/\d{1,2}/\d{2}(\d{2})?$")  # 07/04/25 or 07/04/2025


def _find_header(ws):
    for row in ws.iter_rows():
        labels = [str(c.value).strip().lower() for c in row if c.value is not None]
        if any("date" in v for v in labels) and len(labels) >= 3:
            return row[0].row, [c.value for c in row]
    return None, None


def check(path: str) -> int:
    ws = load_workbook(path).active
    hdr_row, headers = _find_header(ws)
    problems: list[str] = []
    if hdr_row is None:
        print("FAIL: no transaction header row with a 'Date' column found")
        return 1

    headers = [h for h in headers if h is not None]
    date_cols = [i for i, h in enumerate(headers, 1) if "date" in str(h).lower()]
    amt_cols = [i for i, h in enumerate(headers, 1)
                if re.search(r"debit|credit|withdraw|deposit|amount|balance", str(h).lower())]

    date_ok = date_bad = amt_ok = amt_text = 0
    for r in range(hdr_row + 1, ws.max_row + 1):
        for ci in date_cols:
            v = ws.cell(r, ci).value
            if v in (None, ""):
                continue
            if isinstance(v, datetime):
                problems.append(f"row {r} col {ci}: date is an Excel DATE cell "
                                f"({v.date()}) — BUSY wants TEXT like 07/04/25")
                date_bad += 1
            elif isinstance(v, str) and _DATE_TEXT.match(v.strip()):
                date_ok += 1
            else:
                problems.append(f"row {r} col {ci}: date {v!r} not DD/MM/YY text")
                date_bad += 1
        for ci in amt_cols:
            v = ws.cell(r, ci).value
            if v in (None, ""):
                continue
            if isinstance(v, (int, float)):
                amt_ok += 1
            elif isinstance(v, str) and re.search(r"[A-Za-z]", v):
                pass  # e.g. "15,67,705.77 DR" — balance w/ indicator, expected as text
            elif isinstance(v, str):
                amt_text += 1  # numeric-looking but stored as text

    print(f"header row {hdr_row}: {headers}")
    print(f"date columns {date_cols}: {date_ok} OK (text DD/MM/YY), {date_bad} bad")
    print(f"amount columns {amt_cols}: {amt_ok} numeric"
          + (f", {amt_text} stored as text (BUSY may not sum these)" if amt_text else ""))
    if problems:
        print("\nPROBLEMS:")
        for p in problems[:20]:
            print(" -", p)
        return 1
    print("\nPASS: matches the working-file conventions (date=text, amounts=numeric)")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    sys.exit(check(sys.argv[1]))
