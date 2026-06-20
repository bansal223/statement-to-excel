"""Render an extracted Statement into an .xlsx that mirrors the source layout.

Generic / column-driven: whatever header fields and table columns the statement
has (home-loan, current-account, ...) are reproduced as-is. Money columns are
written as numbers with Indian grouping; a balance with a DR/CR suffix is kept as
printed text so the statement reads exactly like the original. Rows whose running
balance fails reconciliation are tinted amber for review.
"""
from __future__ import annotations

import io
import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .vision import Statement

INR_FMT = "#,##,##0.00"
# BUSY imports dates as plain TEXT in DD/MM/YY form (verified against a working
# import file: cells like "07/04/25", type str, General format) — NOT as Excel
# date cells. STATEMENT_DATE_FMT is a Python strftime pattern (e.g. "%d/%m/%Y"
# for a 4-digit year).
DATE_FMT = os.environ.get("STATEMENT_DATE_FMT", "%d/%m/%y")
_HEADER_FONT = Font(bold=True)
_TABLE_HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
_REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
_WIDE_CAP = 60

# Date strings seen across statements: "1 Feb 2026", "01/02/2026", "31-08-2025",
# "2025-08-31". Parsed to a real date so accounting software reads it as a date.
_DATE_PATTERNS = (
    "%d %b %Y", "%d %B %Y", "%d-%b-%Y", "%d-%B-%Y",
    "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
    "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y",
)


def _num(s: str | None) -> float | None:
    """Parse '15,67,705.77 DR' / '22,128.00' / '-26,58,690.03' -> float."""
    if not s:
        return None
    m = re.search(r"-?\d[\d,]*\.?\d*", s)
    return float(m.group(0).replace(",", "")) if m else None


def _parse_date(s: str | None) -> datetime | None:
    """Parse a printed statement date into a datetime, or None if unrecognized."""
    if not s:
        return None
    t = s.strip()
    for fmt in _DATE_PATTERNS:
        try:
            return datetime.strptime(t, fmt)
        except ValueError:
            continue
    return None


def _reconcile(stmt: Statement) -> list[bool]:
    """Return per-row 'looks ok' flags using the running balance, if columns allow."""
    bal_col = next((c for c, r in stmt.roles.items() if r == "balance"), None)
    deb_col = next((c for c, r in stmt.roles.items() if r == "debit"), None)
    cred_col = next((c for c, r in stmt.roles.items() if r == "credit"), None)
    flags = [True] * len(stmt.rows)
    if not bal_col or (not deb_col and not cred_col):
        return flags
    prev = None
    for i, row in enumerate(stmt.rows):
        bal = _num(row.get(bal_col))
        deb = _num(row.get(deb_col)) if deb_col else None
        cred = _num(row.get(cred_col)) if cred_col else None
        amount = deb if deb is not None else cred
        if prev is not None and bal is not None and amount is not None:
            delta = abs(round(bal - prev, 2))
            flags[i] = abs(delta - amount) <= max(1.0, amount * 0.01)
        if bal is not None:
            prev = bal
    return flags


def build_xlsx(stmt: Statement) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"
    flags = _reconcile(stmt)

    r = 1
    # --- header block ---
    for label, value in stmt.header:
        ws.cell(r, 1, f"{label} :").font = _HEADER_FONT
        ws.cell(r, 2, value)
        r += 1
    if stmt.header:
        r += 1  # spacer

    # --- table header ---
    head_row = r
    for ci, name in enumerate(stmt.columns, start=1):
        cell = ws.cell(r, ci, name)
        cell.font = _HEADER_FONT
        cell.fill = _TABLE_HEAD_FILL
    r += 1

    # --- rows ---
    for row, ok in zip(stmt.rows, flags):
        for ci, col in enumerate(stmt.columns, start=1):
            raw = row.get(col)
            role = stmt.roles.get(col)
            cell = ws.cell(r, ci)
            if role in ("debit", "credit") and raw:
                val = _num(raw)
                if val is not None:
                    cell.value = val
                    cell.number_format = INR_FMT
                else:
                    cell.value = raw
            elif role == "balance" and raw:
                # Plain balance -> numeric (matches working BUSY file). A balance
                # carrying a "DR"/"CR" indicator isn't a plain number -> keep text.
                if re.search(r"[A-Za-z]", raw):
                    cell.value = raw
                else:
                    val = _num(raw)
                    if val is not None:
                        cell.value = val
                        cell.number_format = INR_FMT
                    else:
                        cell.value = raw
            elif "date" in col.lower() and raw:
                # BUSY expects the date as TEXT in DD/MM/YY form, not an Excel date
                # cell. Parse the printed value and re-emit it as a plain string.
                d = _parse_date(raw)
                cell.value = d.strftime(DATE_FMT) if d is not None else raw
            elif raw is not None:
                cell.value = raw
            if not ok:
                cell.fill = _REVIEW_FILL
        r += 1

    # --- footer ---
    if stmt.footer:
        r += 1
        ws.cell(r, 1, stmt.footer).font = Font(italic=True)

    _autosize(ws)
    if stmt.columns:
        ws.freeze_panes = ws.cell(head_row + 1, 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _autosize(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(w + 2, _WIDE_CAP)
